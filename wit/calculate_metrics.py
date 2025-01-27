import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

# Read the data from the Historical Data sheet
df = pd.read_excel('travel_market_summary.xlsx', sheet_name='Historical Data (2005-)')

def calculate_metrics_for_group(group):
    """Calculate metrics for each year and market combination"""
    metrics = []
    
    for year in group['Year'].unique():
        year_data = group[group['Year'] == year]
        
        # Calculate total gross bookings
        total_bookings = year_data['Total Market Gross Bookings'].sum()
        
        # Calculate online bookings (sum of Online Supplier-Direct, OTA)
        online_data = year_data[
            year_data['Breakdown'].isin(['Online Supplier-Direct', 'OTA'])
        ]
        online_bookings = online_data['Total Market Gross Bookings'].sum()
        
        # Calculate online penetration
        online_penetration = (online_bookings / total_bookings * 100) if total_bookings > 0 else 0
        
        metrics.append({
            'Year': year,
            'Market': group['Market'].iloc[0],
            'Region': group['Region'].iloc[0],
            'Gross Bookings': total_bookings,
            'Online Bookings': online_bookings,
            'Online Penetration': online_penetration
        })
    
    return pd.DataFrame(metrics)

# Group by Market and Year
grouped = df.groupby('Market')
all_metrics = []

for market, group in grouped:
    market_metrics = calculate_metrics_for_group(group)
    all_metrics.append(market_metrics)

# Combine all metrics
visualization_df = pd.concat(all_metrics, ignore_index=True)

# Sort the data
visualization_df = visualization_df.sort_values(['Year', 'Region', 'Market'])

# Round the numbers for better readability
visualization_df['Online Penetration'] = visualization_df['Online Penetration'].round(2)
visualization_df['Gross Bookings'] = visualization_df['Gross Bookings'].round(2)
visualization_df['Online Bookings'] = visualization_df['Online Bookings'].round(2)

# First save to a temporary file
visualization_df.to_excel('temp_viz.xlsx', sheet_name='Visualization Data', index=False)

# Now read the temporary file with openpyxl
temp_wb = load_workbook('temp_viz.xlsx')
temp_ws = temp_wb.active

# Load the target workbook
wb = load_workbook('travel_market_summary.xlsx')

# Remove the sheet if it exists
if 'Visualization Data' in wb.sheetnames:
    del wb['Visualization Data']

# Create new sheet
wb.create_sheet('Visualization Data')
target_ws = wb['Visualization Data']

# Copy the data
for row in temp_ws.rows:
    for cell in row:
        target_ws[cell.coordinate] = cell.value

# Save the target workbook
wb.save('travel_market_summary.xlsx')

# Clean up temporary file
os.remove('temp_viz.xlsx')

print("Visualization data has been added to 'travel_market_summary.xlsx' with the following metrics:")
print("- Gross Bookings (Total market size)")
print("- Online Bookings (Sum of Online Supplier-Direct and OTA)")
print("- Online Penetration (Online Bookings / Gross Bookings * 100)")
print("\nSample of the data:")
print(visualization_df.head()) 