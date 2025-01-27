import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

# Read the existing data
original_df = pd.read_excel('travel_market_summary.xlsx')

# Convert Year to numeric if it's not already
original_df['Year'] = pd.to_numeric(original_df['Year'])

# Sort by Year to ensure chronological order
original_df = original_df.sort_values(['Market', 'Channel Type', 'Breakdown', 'Year'])

def calculate_growth_rate(start_value, end_value, years_between):
    if start_value <= 0 or end_value <= 0:
        return 0
    return (end_value / start_value) ** (1 / years_between) - 1

def generate_historical_data(group):
    if len(group) < 2:  # Need at least 2 years to calculate growth rate
        return pd.DataFrame()
    
    # Get the earliest two years to calculate growth rate
    earliest_years = group.nsmallest(2, 'Year')
    if len(earliest_years) < 2:
        return pd.DataFrame()
    
    start_value = earliest_years.iloc[0]['Total Market Gross Bookings']
    end_value = earliest_years.iloc[1]['Total Market Gross Bookings']
    years_between = earliest_years.iloc[1]['Year'] - earliest_years.iloc[0]['Year']
    
    growth_rate = calculate_growth_rate(start_value, end_value, years_between)
    
    # Create new rows for 2005-2008
    new_rows = []
    earliest_year = earliest_years.iloc[0]['Year']
    earliest_value = earliest_years.iloc[0]['Total Market Gross Bookings']
    
    # Only generate data for years before 2009
    for year in range(2005, 2009):
        years_projection = earliest_year - year
        new_value = earliest_value / ((1 + growth_rate) ** years_projection)
        
        new_row = earliest_years.iloc[0].copy()
        new_row['Year'] = year
        new_row['Total Market Gross Bookings'] = new_value
        new_rows.append(new_row)
    
    if new_rows:
        return pd.DataFrame(new_rows)
    return pd.DataFrame()

# Group by Market, Channel Type, and Breakdown to maintain consistency
grouped = original_df.groupby(['Market', 'Channel Type', 'Breakdown'])

# Generate historical data (2005-2008)
historical_dfs = []
for name, group in grouped:
    historical_group = generate_historical_data(group)
    if not historical_group.empty:
        historical_dfs.append(historical_group)

# Combine all historical data
if historical_dfs:
    historical_df = pd.concat(historical_dfs, ignore_index=True)
    
    # Combine historical data with original data
    result_df = pd.concat([historical_df, original_df], ignore_index=True)
    
    # Sort the final dataframe
    result_df = result_df.sort_values(['Year', 'Market', 'Channel Type', 'Breakdown'])
    
    # First save the result to a temporary file
    result_df.to_excel('temp_historical.xlsx', sheet_name='Historical Data (2005-)', index=False)
    
    # Now read the temporary file with openpyxl
    temp_wb = load_workbook('temp_historical.xlsx')
    temp_ws = temp_wb.active
    
    # Load the target workbook
    wb = load_workbook('travel_market_summary.xlsx')
    
    # Copy the sheet from temp to target
    if 'Historical Data (2005-)' in wb.sheetnames:
        del wb['Historical Data (2005-)']
    wb.create_sheet('Historical Data (2005-)')
    target_ws = wb['Historical Data (2005-)']
    
    # Copy the data
    for row in temp_ws.rows:
        for cell in row:
            target_ws[cell.coordinate] = cell.value
    
    # Save the target workbook
    wb.save('travel_market_summary.xlsx')
    
    # Clean up temporary file
    os.remove('temp_historical.xlsx')
    
    print("Historical data (2005 onwards) has been added as a new sheet to 'travel_market_summary.xlsx'")
else:
    print("No historical data could be generated") 